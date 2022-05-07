#python excel--openpy
from openpyxl import Workbook
# 有括號不一定是function, 大寫套件是物件
wb = Workbook() #type : Workbook(發明的型別)
# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")